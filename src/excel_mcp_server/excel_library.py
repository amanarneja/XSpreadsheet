"""
Excel Library

A comprehensive library for Excel file manipulation including:
- Reading and writing Excel files
- Worksheet management
- Cell operations
- Formula handling
- Data formatting
- Chart creation
"""

import os
import json
from typing import Any, Dict, List, Optional, Union, Tuple
from pathlib import Path
import logging

import pandas as pd
import openpyxl
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.chart import LineChart, BarChart, PieChart, ScatterChart, Reference
from openpyxl.utils import get_column_letter
import xlsxwriter

logger = logging.getLogger(__name__)


class ExcelLibrary:
    """Main Excel manipulation library"""
    
    def __init__(self):
        """Initialize the Excel library"""
        self.supported_formats = ['.xlsx', '.xls']
        self.chart_types = {
            'line': LineChart,
            'bar': BarChart, 
            'pie': PieChart,
            'scatter': ScatterChart
        }

    def read_excel_file(self, file_path: str, sheet_name: Optional[str] = None, cell_range: Optional[str] = None) -> Dict[str, Any]:
        """
        Read data from an Excel file
        
        Args:
            file_path: Path to the Excel file
            sheet_name: Name of the worksheet (optional)
            cell_range: Cell range to read (optional)
            
        Returns:
            Dictionary containing the data and metadata
        """
        try:
            if not os.path.exists(file_path):
                raise FileNotFoundError(f"File not found: {file_path}")
            
            # Load workbook
            workbook = load_workbook(file_path, data_only=True)
            
            # Get worksheet
            if sheet_name:
                if sheet_name not in workbook.sheetnames:
                    raise ValueError(f"Worksheet '{sheet_name}' not found")
                worksheet = workbook[sheet_name]
            else:
                worksheet = workbook.active
                sheet_name = worksheet.title
            
            # Read data
            if cell_range:
                cells = worksheet[cell_range]
                if isinstance(cells, tuple):
                    # Multiple rows
                    data = [[cell.value for cell in row] for row in cells]
                else:
                    # Single row or cell
                    data = [[cell.value for cell in cells]]
            else:
                # Read all data
                data = []
                for row in worksheet.iter_rows(values_only=True):
                    if any(cell is not None for cell in row):  # Skip empty rows
                        data.append(list(row))
            
            workbook.close()
            
            return {
                "success": True,
                "data": data,
                "sheet_name": sheet_name,
                "range": cell_range or f"A1:{get_column_letter(worksheet.max_column)}{worksheet.max_row}",
                "rows": len(data),
                "columns": len(data[0]) if data else 0
            }
            
        except Exception as e:
            logger.error(f"Error reading Excel file: {str(e)}")
            return {
                "success": False,
                "error": str(e),
                "data": []
            }

    def write_excel_file(self, file_path: str, data: List[List[Any]], sheet_name: Optional[str] = None, headers: Optional[List[str]] = None) -> Dict[str, Any]:
        """
        Write data to an Excel file
        
        Args:
            file_path: Path where to save the Excel file
            data: 2D array of data to write
            sheet_name: Name of the worksheet (optional)
            headers: Column headers (optional)
            
        Returns:
            Dictionary containing operation result
        """
        try:
            if not data:
                raise ValueError("No data provided to write")
            
            # Create or load workbook
            if os.path.exists(file_path):
                workbook = load_workbook(file_path)
            else:
                workbook = Workbook()
                # Remove default sheet if we're naming our sheet
                if sheet_name and 'Sheet' in workbook.sheetnames:
                    workbook.remove(workbook['Sheet'])
            
            # Create or get worksheet
            if sheet_name:
                if sheet_name in workbook.sheetnames:
                    worksheet = workbook[sheet_name]
                    # Clear existing data
                    worksheet.delete_rows(1, worksheet.max_row)
                else:
                    worksheet = workbook.create_sheet(sheet_name)
            else:
                worksheet = workbook.active
                sheet_name = worksheet.title
            
            # Write headers if provided
            row_offset = 0
            if headers:
                for col, header in enumerate(headers, 1):
                    worksheet.cell(row=1, column=col, value=header)
                row_offset = 1
            
            # Write data
            for row_idx, row_data in enumerate(data, 1 + row_offset):
                for col_idx, value in enumerate(row_data, 1):
                    worksheet.cell(row=row_idx, column=col_idx, value=value)
            
            # Save workbook
            workbook.save(file_path)
            workbook.close()
            
            return {
                "success": True,
                "file_path": file_path,
                "sheet_name": sheet_name,
                "rows_written": len(data),
                "columns_written": len(data[0]) if data else 0,
                "headers_added": bool(headers)
            }
            
        except Exception as e:
            logger.error(f"Error writing Excel file: {str(e)}")
            return {
                "success": False,
                "error": str(e)
            }

    def get_worksheet_info(self, file_path: str) -> Dict[str, Any]:
        """
        Get information about worksheets in an Excel file
        
        Args:
            file_path: Path to the Excel file
            
        Returns:
            Dictionary containing worksheet information
        """
        try:
            if not os.path.exists(file_path):
                raise FileNotFoundError(f"File not found: {file_path}")
            
            workbook = load_workbook(file_path, data_only=True)
            
            worksheets = []
            for sheet_name in workbook.sheetnames:
                worksheet = workbook[sheet_name]
                worksheets.append({
                    "name": sheet_name,
                    "max_row": worksheet.max_row,
                    "max_column": worksheet.max_column,
                    "max_column_letter": get_column_letter(worksheet.max_column),
                    "is_active": worksheet == workbook.active
                })
            
            workbook.close()
            
            return {
                "success": True,
                "file_path": file_path,
                "total_worksheets": len(worksheets),
                "worksheets": worksheets
            }
            
        except Exception as e:
            logger.error(f"Error getting worksheet info: {str(e)}")
            return {
                "success": False,
                "error": str(e)
            }

    def add_worksheet(self, file_path: str, sheet_name: str) -> Dict[str, Any]:
        """
        Add a new worksheet to an existing Excel file
        
        Args:
            file_path: Path to the Excel file
            sheet_name: Name of the new worksheet
            
        Returns:
            Dictionary containing operation result
        """
        try:
            if not os.path.exists(file_path):
                raise FileNotFoundError(f"File not found: {file_path}")
            
            workbook = load_workbook(file_path)
            
            if sheet_name in workbook.sheetnames:
                raise ValueError(f"Worksheet '{sheet_name}' already exists")
            
            worksheet = workbook.create_sheet(sheet_name)
            workbook.save(file_path)
            workbook.close()
            
            return {
                "success": True,
                "file_path": file_path,
                "sheet_name": sheet_name,
                "message": f"Worksheet '{sheet_name}' added successfully"
            }
            
        except Exception as e:
            logger.error(f"Error adding worksheet: {str(e)}")
            return {
                "success": False,
                "error": str(e)
            }

    def update_cell(self, file_path: str, sheet_name: str, cell: str, value: Any) -> Dict[str, Any]:
        """
        Update a specific cell in an Excel file
        
        Args:
            file_path: Path to the Excel file
            sheet_name: Name of the worksheet
            cell: Cell reference (e.g., 'A1')
            value: Value to set in the cell
            
        Returns:
            Dictionary containing operation result
        """
        try:
            if not os.path.exists(file_path):
                raise FileNotFoundError(f"File not found: {file_path}")
            
            workbook = load_workbook(file_path)
            
            if sheet_name not in workbook.sheetnames:
                raise ValueError(f"Worksheet '{sheet_name}' not found")
            
            worksheet = workbook[sheet_name]
            worksheet[cell] = value
            
            workbook.save(file_path)
            workbook.close()
            
            return {
                "success": True,
                "file_path": file_path,
                "sheet_name": sheet_name,
                "cell": cell,
                "value": value,
                "message": f"Cell {cell} updated successfully"
            }
            
        except Exception as e:
            logger.error(f"Error updating cell: {str(e)}")
            return {
                "success": False,
                "error": str(e)
            }

    def apply_formula(self, file_path: str, sheet_name: str, cell: str, formula: str) -> Dict[str, Any]:
        """
        Apply a formula to a cell in an Excel file
        
        Args:
            file_path: Path to the Excel file
            sheet_name: Name of the worksheet
            cell: Cell reference (e.g., 'A1')
            formula: Excel formula (e.g., '=SUM(A1:A10)')
            
        Returns:
            Dictionary containing operation result
        """
        try:
            if not os.path.exists(file_path):
                raise FileNotFoundError(f"File not found: {file_path}")
            
            if not formula.startswith('='):
                formula = '=' + formula
            
            workbook = load_workbook(file_path)
            
            if sheet_name not in workbook.sheetnames:
                raise ValueError(f"Worksheet '{sheet_name}' not found")
            
            worksheet = workbook[sheet_name]
            worksheet[cell] = formula
            
            workbook.save(file_path)
            workbook.close()
            
            return {
                "success": True,
                "file_path": file_path,
                "sheet_name": sheet_name,
                "cell": cell,
                "formula": formula,
                "message": f"Formula applied to cell {cell} successfully"
            }
            
        except Exception as e:
            logger.error(f"Error applying formula: {str(e)}")
            return {
                "success": False,
                "error": str(e)
            }

    def format_cells(self, file_path: str, sheet_name: str, cell_range: str, format_options: Dict[str, Any]) -> Dict[str, Any]:
        """
        Apply formatting to cells in an Excel file
        
        Args:
            file_path: Path to the Excel file
            sheet_name: Name of the worksheet
            cell_range: Cell range to format (e.g., 'A1:C10')
            format_options: Formatting options
            
        Returns:
            Dictionary containing operation result
        """
        try:
            if not os.path.exists(file_path):
                raise FileNotFoundError(f"File not found: {file_path}")
            
            workbook = load_workbook(file_path)
            
            if sheet_name not in workbook.sheetnames:
                raise ValueError(f"Worksheet '{sheet_name}' not found")
            
            worksheet = workbook[sheet_name]
            
            # Create font style
            font_kwargs = {}
            if format_options.get('bold'):
                font_kwargs['bold'] = True
            if format_options.get('italic'):
                font_kwargs['italic'] = True
            if format_options.get('font_size'):
                font_kwargs['size'] = format_options['font_size']
            if format_options.get('font_color'):
                font_kwargs['color'] = format_options['font_color']
            
            font = Font(**font_kwargs) if font_kwargs else None
            
            # Create fill style
            fill = None
            if format_options.get('background_color'):
                fill = PatternFill(start_color=format_options['background_color'], 
                                 end_color=format_options['background_color'], 
                                 fill_type='solid')
            
            # Apply formatting to range
            for row in worksheet[cell_range]:
                for cell in row:
                    if font:
                        cell.font = font
                    if fill:
                        cell.fill = fill
                    if format_options.get('number_format'):
                        cell.number_format = format_options['number_format']
            
            workbook.save(file_path)
            workbook.close()
            
            return {
                "success": True,
                "file_path": file_path,
                "sheet_name": sheet_name,
                "range": cell_range,
                "format_options": format_options,
                "message": f"Formatting applied to range {cell_range} successfully"
            }
            
        except Exception as e:
            logger.error(f"Error formatting cells: {str(e)}")
            return {
                "success": False,
                "error": str(e)
            }

    def create_chart(self, file_path: str, sheet_name: str, data_range: str, chart_type: str, title: Optional[str] = None, position: Optional[str] = None) -> Dict[str, Any]:
        """
        Create a chart in an Excel file
        
        Args:
            file_path: Path to the Excel file
            sheet_name: Name of the worksheet
            data_range: Data range for the chart (e.g., 'A1:B10')
            chart_type: Type of chart (line, bar, pie, scatter)
            title: Chart title (optional)
            position: Chart position (optional)
            
        Returns:
            Dictionary containing operation result
        """
        try:
            if not os.path.exists(file_path):
                raise FileNotFoundError(f"File not found: {file_path}")
            
            if chart_type not in self.chart_types:
                raise ValueError(f"Unsupported chart type: {chart_type}")
            
            workbook = load_workbook(file_path)
            
            if sheet_name not in workbook.sheetnames:
                raise ValueError(f"Worksheet '{sheet_name}' not found")
            
            worksheet = workbook[sheet_name]
            
            # Create chart
            chart = self.chart_types[chart_type]()
            if title:
                chart.title = title
            
            # Add data to chart
            data = Reference(worksheet, range_string=f"{sheet_name}!{data_range}")
            chart.add_data(data, titles_from_data=True)
            
            # Position chart
            if position:
                chart.anchor = position
            else:
                # Default position
                chart.anchor = "E5"
            
            worksheet.add_chart(chart)
            
            workbook.save(file_path)
            workbook.close()
            
            return {
                "success": True,
                "file_path": file_path,
                "sheet_name": sheet_name,
                "data_range": data_range,
                "chart_type": chart_type,
                "title": title,
                "position": position or "E5",
                "message": f"{chart_type.title()} chart created successfully"
            }
            
        except Exception as e:
            logger.error(f"Error creating chart: {str(e)}")
            return {
                "success": False,
                "error": str(e)
            }

    def convert_to_pandas(self, file_path: str, sheet_name: Optional[str] = None) -> pd.DataFrame:
        """
        Convert Excel data to a pandas DataFrame
        
        Args:
            file_path: Path to the Excel file
            sheet_name: Name of the worksheet (optional)
            
        Returns:
            pandas DataFrame
        """
        try:
            df = pd.read_excel(file_path, sheet_name=sheet_name)
            return df
        except Exception as e:
            logger.error(f"Error converting to pandas: {str(e)}")
            return pd.DataFrame()

    def from_pandas(self, df: pd.DataFrame, file_path: str, sheet_name: Optional[str] = None) -> Dict[str, Any]:
        """
        Write pandas DataFrame to Excel file
        
        Args:
            df: pandas DataFrame
            file_path: Path where to save the Excel file
            sheet_name: Name of the worksheet (optional)
            
        Returns:
            Dictionary containing operation result
        """
        try:
            df.to_excel(file_path, sheet_name=sheet_name or 'Sheet1', index=False)
            
            return {
                "success": True,
                "file_path": file_path,
                "sheet_name": sheet_name or 'Sheet1',
                "rows": len(df),
                "columns": len(df.columns),
                "message": "DataFrame written to Excel successfully"
            }
            
        except Exception as e:
            logger.error(f"Error writing pandas to Excel: {str(e)}")
            return {
                "success": False,
                "error": str(e)
            }
